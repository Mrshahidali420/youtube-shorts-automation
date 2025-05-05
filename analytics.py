#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Performance Analytics Module

This module provides enhanced analytics for YouTube channel performance,
including trend analysis, content performance correlation, and visualization.

Copyright (c) 2023-2025 Shahid Ali
License: MIT License
GitHub: https://github.com/Mrshahidali420/youtube-shorts-automation
Version: 1.5.0
"""

import os
import json
import time
import datetime
import matplotlib.pyplot as plt
import numpy as np
from typing import Dict, List, Tuple, Any, Optional
from collections import defaultdict
import pandas as pd
import traceback

# Try to import YouTube API libraries
try:
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    YOUTUBE_API_AVAILABLE = True
except ImportError:
    print("Warning: Google API libraries not found. Some analytics features will be limited.")
    YOUTUBE_API_AVAILABLE = False

# --- Constants ---
ANALYTICS_FOLDER = "analytics_data"
PERFORMANCE_HISTORY_FILE = "performance_history.json"
ANALYTICS_REPORT_FOLDER = "analytics_reports"
CORRELATION_ANALYSIS_FILE = "correlation_analysis.json"
AUDIENCE_INSIGHTS_FILE = "audience_insights.json"
CONFIG_FILENAME = "config.txt"  # To read API keys

# --- Setup ---
script_directory = os.path.dirname(os.path.abspath(__file__))
analytics_folder = os.path.join(script_directory, ANALYTICS_FOLDER)
analytics_report_folder = os.path.join(analytics_folder, ANALYTICS_REPORT_FOLDER)
performance_history_path = os.path.join(analytics_folder, PERFORMANCE_HISTORY_FILE)
correlation_analysis_path = os.path.join(analytics_folder, CORRELATION_ANALYSIS_FILE)
audience_insights_path = os.path.join(analytics_folder, AUDIENCE_INSIGHTS_FILE)
config_file_path = os.path.join(script_directory, CONFIG_FILENAME)

# Create necessary folders
os.makedirs(analytics_folder, exist_ok=True)
os.makedirs(analytics_report_folder, exist_ok=True)

# --- Logging Functions ---
def log_info(msg: str) -> None:
    """Log an informational message."""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] INFO: {msg}")

def log_error(msg: str, include_traceback: bool = False) -> None:
    """Log an error message with optional traceback."""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] ERROR: {msg}")
    if include_traceback:
        traceback.print_exc()

# --- Configuration Loading ---
def load_config() -> Dict[str, str]:
    """Load configuration from config.txt file."""
    config = {}
    try:
        with open(config_file_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and "=" in line:
                    key, value = line.split("=", 1)
                    config[key.strip()] = value.strip()
        return config
    except Exception as e:
        log_error(f"Error loading config: {e}", include_traceback=True)
        return {}

# --- Data Loading Functions ---
def load_performance_history() -> Dict[str, Any]:
    """Load historical performance data from JSON file."""
    if not os.path.exists(performance_history_path):
        return {"videos": [], "last_updated": "", "metrics": {}}

    try:
        with open(performance_history_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log_error(f"Error loading performance history: {e}", include_traceback=True)
        return {"videos": [], "last_updated": "", "metrics": {}}

def save_performance_history(data: Dict[str, Any]) -> None:
    """Save performance history data to JSON file."""
    try:
        data["last_updated"] = datetime.datetime.now().isoformat()
        with open(performance_history_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        log_info(f"Performance history saved to {performance_history_path}")
    except Exception as e:
        log_error(f"Error saving performance history: {e}", include_traceback=True)

# --- YouTube API Functions ---
def get_youtube_api_client() -> Optional[Any]:
    """Initialize and return a YouTube API client."""
    if not YOUTUBE_API_AVAILABLE:
        log_error("YouTube API libraries not available.")
        return None

    config = load_config()
    api_key = config.get("API_KEY") or config.get("YOUTUBE_API_KEY")

    if not api_key:
        log_error("API_KEY not found in config.txt")
        return None

    try:
        youtube = build("youtube", "v3", developerKey=api_key)
        return youtube
    except Exception as e:
        log_error(f"Error initializing YouTube API client: {e}", include_traceback=True)
        return None

def fetch_video_statistics(video_ids: List[str]) -> Dict[str, Dict[str, Any]]:
    """Fetch current statistics for a list of video IDs using YouTube API."""
    if not video_ids:
        return {}

    youtube = get_youtube_api_client()
    if not youtube:
        return {}

    results = {}
    # Process in batches of 50 (API limit)
    for i in range(0, len(video_ids), 50):
        batch = video_ids[i:i+50]
        try:
            response = youtube.videos().list(
                part="statistics,snippet",
                id=",".join(batch)
            ).execute()

            for item in response.get("items", []):
                video_id = item["id"]
                stats = item.get("statistics", {})
                snippet = item.get("snippet", {})

                results[video_id] = {
                    "views": int(stats.get("viewCount", 0)),
                    "likes": int(stats.get("likeCount", 0)),
                    "comments": int(stats.get("commentCount", 0)),
                    "title": snippet.get("title", ""),
                    "published_at": snippet.get("publishedAt", ""),
                    "timestamp": datetime.datetime.now().isoformat()
                }

            # Respect API quota by adding a small delay between batch requests
            if i + 50 < len(video_ids):
                time.sleep(1)

        except HttpError as e:
            log_error(f"YouTube API error: {e}", include_traceback=True)
        except Exception as e:
            log_error(f"Error fetching video statistics: {e}", include_traceback=True)

    return results

# --- Analytics Functions ---
def update_performance_history(excel_file_path: str) -> None:
    """
    Update performance history by reading the Excel file and fetching latest stats.

    Args:
        excel_file_path: Path to the Excel file containing uploaded videos data
    """
    try:
        # Import here to avoid circular imports
        import pandas as pd
        from openpyxl import load_workbook

        # Load existing history
        history = load_performance_history()

        # Read Excel file
        wb = load_workbook(excel_file_path, read_only=True)
        if "Uploaded" not in wb.sheetnames:
            log_error("Uploaded sheet not found in Excel file")
            return

        # Convert sheet to DataFrame
        sheet = wb["Uploaded"]
        data = []
        headers = []

        # Get headers from first row
        for cell in sheet[1]:
            headers.append(cell.value)

        # Get data from remaining rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                data.append(row)

        df = pd.DataFrame(data, columns=headers)

        # Clean column names
        df.columns = [str(col).lower().strip() for col in df.columns]

        # Find the YouTube ID column
        yt_id_col = next((col for col in df.columns if 'youtube' in col and 'id' in col), None)
        if not yt_id_col:
            log_error("YouTube ID column not found in Excel file")
            return

        # Get list of video IDs
        video_ids = df[yt_id_col].dropna().tolist()

        # Fetch latest statistics from YouTube API
        latest_stats = fetch_video_statistics(video_ids)

        # Update history
        existing_ids = {v["video_id"]: i for i, v in enumerate(history["videos"])}

        for video_id, stats in latest_stats.items():
            if video_id in existing_ids:
                # Update existing entry
                idx = existing_ids[video_id]
                history["videos"][idx]["stats_history"].append(stats)
            else:
                # Create new entry
                video_row = df[df[yt_id_col] == video_id].iloc[0]

                # Find title column
                title_col = next((col for col in df.columns if 'title' in col), None)
                title = video_row[title_col] if title_col else "Unknown Title"

                # Find upload date column
                upload_col = next((col for col in df.columns if 'upload' in col or 'timestamp' in col), None)
                upload_date = video_row[upload_col] if upload_col else ""

                new_entry = {
                    "video_id": video_id,
                    "title": title,
                    "upload_date": upload_date,
                    "stats_history": [stats]
                }
                history["videos"].append(new_entry)

        # Save updated history
        save_performance_history(history)
        log_info(f"Updated performance history for {len(latest_stats)} videos")

    except Exception as e:
        log_error(f"Error updating performance history: {e}", include_traceback=True)

def calculate_growth_metrics() -> Dict[str, Any]:
    """Calculate growth metrics from performance history."""
    history = load_performance_history()
    metrics = {
        "total_videos": len(history["videos"]),
        "total_views": 0,
        "total_likes": 0,
        "total_comments": 0,
        "avg_views_per_video": 0,
        "avg_likes_per_video": 0,
        "avg_comments_per_video": 0,
        "engagement_rate": 0,  # (likes + comments) / views
        "growth_rate_7d": 0,   # 7-day view growth rate
        "growth_rate_30d": 0,  # 30-day view growth rate
        "top_performing_videos": [],
        "recent_performance": []
    }

    if not history["videos"]:
        return metrics

    # Calculate total metrics
    for video in history["videos"]:
        if not video["stats_history"]:
            continue

        latest_stats = video["stats_history"][-1]
        metrics["total_views"] += latest_stats.get("views", 0)
        metrics["total_likes"] += latest_stats.get("likes", 0)
        metrics["total_comments"] += latest_stats.get("comments", 0)

    # Calculate averages
    if metrics["total_videos"] > 0:
        metrics["avg_views_per_video"] = metrics["total_views"] / metrics["total_videos"]
        metrics["avg_likes_per_video"] = metrics["total_likes"] / metrics["total_videos"]
        metrics["avg_comments_per_video"] = metrics["total_comments"] / metrics["total_videos"]

    # Calculate engagement rate
    if metrics["total_views"] > 0:
        metrics["engagement_rate"] = (metrics["total_likes"] + metrics["total_comments"]) / metrics["total_views"] * 100

    # Find top performing videos (by views)
    videos_with_stats = []
    for video in history["videos"]:
        if video["stats_history"]:
            latest_stats = video["stats_history"][-1]
            videos_with_stats.append({
                "video_id": video["video_id"],
                "title": video["title"],
                "views": latest_stats.get("views", 0),
                "likes": latest_stats.get("likes", 0),
                "comments": latest_stats.get("comments", 0),
                "engagement_rate": (latest_stats.get("likes", 0) + latest_stats.get("comments", 0)) / max(1, latest_stats.get("views", 1)) * 100
            })

    # Sort by views and get top 10
    top_by_views = sorted(videos_with_stats, key=lambda x: x["views"], reverse=True)[:10]
    metrics["top_performing_videos"] = top_by_views

    # Calculate growth rates (if we have enough history)
    # This is a simplified calculation - a more sophisticated version would use daily data
    now = datetime.datetime.now()
    seven_days_ago = now - datetime.timedelta(days=7)
    thirty_days_ago = now - datetime.timedelta(days=30)

    views_now = metrics["total_views"]
    views_7d_ago = 0
    views_30d_ago = 0

    for video in history["videos"]:
        for stats in video["stats_history"]:
            stats_time = datetime.datetime.fromisoformat(stats["timestamp"].split("+")[0])
            if abs((seven_days_ago - stats_time).days) < 1:  # Close to 7 days ago
                views_7d_ago += stats.get("views", 0)
            if abs((thirty_days_ago - stats_time).days) < 1:  # Close to 30 days ago
                views_30d_ago += stats.get("views", 0)

    # Calculate growth rates
    if views_7d_ago > 0:
        metrics["growth_rate_7d"] = ((views_now - views_7d_ago) / views_7d_ago) * 100

    if views_30d_ago > 0:
        metrics["growth_rate_30d"] = ((views_now - views_30d_ago) / views_30d_ago) * 100

    return metrics

def generate_performance_report(output_path: Optional[str] = None) -> str:
    """
    Generate a comprehensive performance report in HTML format.

    Args:
        output_path: Optional path to save the report. If None, saves to default location.

    Returns:
        Path to the generated report file
    """
    if output_path is None:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(analytics_report_folder, f"performance_report_{timestamp}.html")

    # Calculate metrics
    metrics = calculate_growth_metrics()
    history = load_performance_history()

    # Generate HTML report
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>YouTube Channel Performance Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1, h2, h3 {{ color: #c00; }}
            .metric-card {{ background: #f5f5f5; padding: 15px; margin: 10px 0; border-radius: 5px; }}
            .metric-value {{ font-size: 24px; font-weight: bold; color: #333; }}
            .metric-label {{ font-size: 14px; color: #666; }}
            .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 15px; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th, td {{ text-align: left; padding: 8px; border-bottom: 1px solid #ddd; }}
            th {{ background-color: #f2f2f2; }}
            tr:hover {{ background-color: #f5f5f5; }}
        </style>
    </head>
    <body>
        <h1>YouTube Channel Performance Report</h1>
        <p>Generated on {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>

        <h2>Channel Overview</h2>
        <div class="metric-grid">
            <div class="metric-card">
                <div class="metric-value">{metrics["total_videos"]}</div>
                <div class="metric-label">Total Videos</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics["total_views"]:,}</div>
                <div class="metric-label">Total Views</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics["total_likes"]:,}</div>
                <div class="metric-label">Total Likes</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics["total_comments"]:,}</div>
                <div class="metric-label">Total Comments</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics["avg_views_per_video"]:,.1f}</div>
                <div class="metric-label">Avg Views Per Video</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics["engagement_rate"]:.2f}%</div>
                <div class="metric-label">Engagement Rate</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics["growth_rate_7d"]:.1f}%</div>
                <div class="metric-label">7-Day Growth Rate</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics["growth_rate_30d"]:.1f}%</div>
                <div class="metric-label">30-Day Growth Rate</div>
            </div>
        </div>

        <h2>Top Performing Videos</h2>
        <table>
            <tr>
                <th>Title</th>
                <th>Views</th>
                <th>Likes</th>
                <th>Comments</th>
                <th>Engagement Rate</th>
            </tr>
    """

    # Add top videos to the table
    for video in metrics["top_performing_videos"]:
        html += f"""
            <tr>
                <td><a href="https://www.youtube.com/watch?v={video["video_id"]}" target="_blank">{video["title"]}</a></td>
                <td>{video["views"]:,}</td>
                <td>{video["likes"]:,}</td>
                <td>{video["comments"]:,}</td>
                <td>{video["engagement_rate"]:.2f}%</td>
            </tr>
        """

    html += """
        </table>

        <h2>Performance Trends</h2>
        <p>This section would include charts and graphs of performance trends over time.</p>

        <h2>Content Analysis</h2>
        <p>This section would include analysis of what types of content perform best.</p>

        <h2>Recommendations</h2>
        <p>This section would include recommendations for improving channel performance.</p>
    </body>
    </html>
    """

    # Save the report
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        log_info(f"Performance report generated: {output_path}")
        return output_path
    except Exception as e:
        log_error(f"Error saving performance report: {e}", include_traceback=True)
        return ""

def analyze_content_correlation() -> Dict[str, Any]:
    """
    Analyze correlation between video characteristics and performance.

    Returns:
        Dictionary of correlation findings
    """
    history = load_performance_history()

    if not history["videos"]:
        return {"status": "No data available"}

    # Extract features for analysis
    video_data = []
    for video in history["videos"]:
        if not video["stats_history"]:
            continue

        latest_stats = video["stats_history"][-1]
        title = video["title"]

        # Extract basic features
        features = {
            "video_id": video["video_id"],
            "title": title,
            "title_length": len(title),
            "has_number_in_title": any(c.isdigit() for c in title),
            "has_emoji_in_title": any(c in title for c in "🔥💪👀😱🤯"),
            "views": latest_stats.get("views", 0),
            "likes": latest_stats.get("likes", 0),
            "comments": latest_stats.get("comments", 0),
            "engagement_rate": (latest_stats.get("likes", 0) + latest_stats.get("comments", 0)) / max(1, latest_stats.get("views", 1)) * 100
        }

        # Add upload day and time features
        try:
            if video.get("upload_date"):
                upload_date = datetime.datetime.fromisoformat(video["upload_date"].split("+")[0])
                features["upload_day"] = upload_date.strftime("%A")  # Day of week
                features["upload_hour"] = upload_date.hour
        except:
            pass

        video_data.append(features)

    # Convert to DataFrame for analysis
    try:
        df = pd.DataFrame(video_data)

        # Calculate correlations
        correlations = {}

        # Title length vs. views
        if "title_length" in df.columns and "views" in df.columns:
            correlations["title_length_vs_views"] = df[["title_length", "views"]].corr().iloc[0, 1]

        # Numbers in title vs. views
        if "has_number_in_title" in df.columns and "views" in df.columns:
            avg_views_with_numbers = df[df["has_number_in_title"]]["views"].mean()
            avg_views_without_numbers = df[~df["has_number_in_title"]]["views"].mean()
            correlations["numbers_in_title_effect"] = {
                "avg_views_with_numbers": avg_views_with_numbers,
                "avg_views_without_numbers": avg_views_without_numbers,
                "percentage_difference": ((avg_views_with_numbers / avg_views_without_numbers) - 1) * 100 if avg_views_without_numbers > 0 else 0
            }

        # Emoji in title vs. views
        if "has_emoji_in_title" in df.columns and "views" in df.columns:
            avg_views_with_emoji = df[df["has_emoji_in_title"]]["views"].mean()
            avg_views_without_emoji = df[~df["has_emoji_in_title"]]["views"].mean()
            correlations["emoji_in_title_effect"] = {
                "avg_views_with_emoji": avg_views_with_emoji,
                "avg_views_without_emoji": avg_views_without_emoji,
                "percentage_difference": ((avg_views_with_emoji / avg_views_without_emoji) - 1) * 100 if avg_views_without_emoji > 0 else 0
            }

        # Upload day vs. views
        if "upload_day" in df.columns and "views" in df.columns:
            day_performance = df.groupby("upload_day")["views"].mean().to_dict()
            best_day = max(day_performance.items(), key=lambda x: x[1])
            correlations["upload_day_performance"] = {
                "day_averages": day_performance,
                "best_day": best_day[0],
                "best_day_avg_views": best_day[1]
            }

        # Upload hour vs. views
        if "upload_hour" in df.columns and "views" in df.columns:
            hour_performance = df.groupby("upload_hour")["views"].mean().to_dict()
            best_hour = max(hour_performance.items(), key=lambda x: x[1])
            correlations["upload_hour_performance"] = {
                "hour_averages": hour_performance,
                "best_hour": best_hour[0],
                "best_hour_avg_views": best_hour[1]
            }

        # Save correlation analysis
        try:
            with open(correlation_analysis_path, "w", encoding="utf-8") as f:
                json.dump(correlations, f, ensure_ascii=False, indent=4)
            log_info(f"Correlation analysis saved to {correlation_analysis_path}")
        except Exception as e:
            log_error(f"Error saving correlation analysis: {e}", include_traceback=True)

        return correlations

    except Exception as e:
        log_error(f"Error analyzing content correlation: {e}", include_traceback=True)
        return {"status": f"Error: {str(e)}"}

# --- Main Functions ---
def run_analytics_update(excel_file_path: str) -> None:
    """
    Run a complete analytics update.

    Args:
        excel_file_path: Path to the Excel file containing uploaded videos data
    """
    log_info("Starting analytics update...")

    # Update performance history
    update_performance_history(excel_file_path)

    # Calculate metrics
    metrics = calculate_growth_metrics()
    log_info(f"Channel metrics calculated: {metrics['total_videos']} videos, {metrics['total_views']:,} views")

    # Analyze content correlation
    correlations = analyze_content_correlation()
    log_info("Content correlation analysis completed")

    # Generate performance report
    report_path = generate_performance_report()
    if report_path:
        log_info(f"Performance report generated: {report_path}")

    log_info("Analytics update completed")

def get_optimization_recommendations() -> Dict[str, Any]:
    """
    Generate optimization recommendations based on analytics data.

    Returns:
        Dictionary of recommendations
    """
    # Load correlation analysis
    try:
        with open(correlation_analysis_path, "r", encoding="utf-8") as f:
            correlations = json.load(f)
    except:
        correlations = {}

    recommendations = {
        "title_recommendations": [],
        "upload_timing_recommendations": [],
        "content_recommendations": []
    }

    # Title recommendations
    if "title_length_vs_views" in correlations:
        corr = correlations["title_length_vs_views"]
        if corr > 0.3:
            recommendations["title_recommendations"].append("Longer titles tend to perform better. Consider using more descriptive titles.")
        elif corr < -0.3:
            recommendations["title_recommendations"].append("Shorter titles tend to perform better. Consider more concise titles.")

    if "numbers_in_title_effect" in correlations:
        effect = correlations["numbers_in_title_effect"]
        if effect.get("percentage_difference", 0) > 10:
            recommendations["title_recommendations"].append("Titles with numbers perform better. Consider including numbers in your titles.")
        elif effect.get("percentage_difference", 0) < -10:
            recommendations["title_recommendations"].append("Titles without numbers perform better. Consider avoiding numbers in your titles.")

    if "emoji_in_title_effect" in correlations:
        effect = correlations["emoji_in_title_effect"]
        if effect.get("percentage_difference", 0) > 10:
            recommendations["title_recommendations"].append("Titles with emojis perform better. Consider including relevant emojis in your titles.")
        elif effect.get("percentage_difference", 0) < -10:
            recommendations["title_recommendations"].append("Titles without emojis perform better. Consider avoiding emojis in your titles.")

    # Upload timing recommendations
    if "upload_day_performance" in correlations:
        day_perf = correlations["upload_day_performance"]
        if day_perf.get("best_day"):
            recommendations["upload_timing_recommendations"].append(f"Videos uploaded on {day_perf['best_day']} tend to perform best. Consider scheduling more uploads on this day.")

    if "upload_hour_performance" in correlations:
        hour_perf = correlations["upload_hour_performance"]
        if hour_perf.get("best_hour") is not None:
            best_hour = hour_perf["best_hour"]
            am_pm = "AM" if best_hour < 12 else "PM"
            hour_12 = best_hour % 12
            if hour_12 == 0:
                hour_12 = 12
            recommendations["upload_timing_recommendations"].append(f"Videos uploaded around {hour_12} {am_pm} tend to perform best. Consider scheduling uploads around this time.")

    # Content recommendations based on top performing videos
    metrics = calculate_growth_metrics()
    top_videos = metrics.get("top_performing_videos", [])

    if top_videos:
        # Analyze titles of top videos for common themes
        titles = [video["title"].lower() for video in top_videos]
        common_words = defaultdict(int)

        for title in titles:
            words = title.split()
            for word in words:
                if len(word) > 3:  # Skip short words
                    common_words[word] += 1

        # Find most common words
        top_words = sorted(common_words.items(), key=lambda x: x[1], reverse=True)[:5]
        if top_words:
            word_list = ", ".join([word for word, count in top_words])
            recommendations["content_recommendations"].append(f"Your top performing videos often contain these keywords: {word_list}. Consider creating more content around these themes.")

    return recommendations

# --- Main Execution ---
if __name__ == "__main__":
    # This code runs when the script is executed directly
    import sys

    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        run_analytics_update(excel_path)
    else:
        print("Usage: python analytics.py <path_to_excel_file>")
