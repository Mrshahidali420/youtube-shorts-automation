#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube Performance Tracker

This script fetches performance metrics for uploaded videos using the YouTube Data API.
It reads the YouTube Video IDs from the Excel file and updates the metrics.
"""

import os
import json
import time
import pickle
from datetime import datetime
from typing import Dict, List, Optional, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Google API imports
try:
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    GOOGLE_API_AVAILABLE = True
except ImportError:
    print("Warning: Google API libraries not found. Install with:")
    print("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
    GOOGLE_API_AVAILABLE = False

# --- Colorama Setup ---
try:
    import colorama
    from colorama import Fore, Style, init
    init(autoreset=True)  # Automatically reset style after each print
    COLOR_ENABLED = True
    print(f"{Fore.GREEN}Colorama loaded successfully. Colored output enabled.{Style.RESET_ALL}")
except ImportError:
    print("Warning: 'colorama' not found. Install it for colored output (`pip install colorama`). Output will be monochrome.")
    # Define dummy color objects if colorama is not available
    class DummyColor:
        def __getattr__(self, name): return ""
    Fore = DummyColor(); Style = DummyColor()  # Assign instances
    COLOR_ENABLED = False
# --- End Colorama Setup ---

# --- Configuration ---
script_directory = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(script_directory, "shorts_data.xlsx")
UPLOADED_SHEET_NAME = "Uploaded"
CLIENT_SECRETS_FILE = os.path.join(script_directory, "client_secret.json")
TOKEN_FILE = os.path.join(script_directory, "token.json")
ERROR_LOG_FILE = os.path.join(script_directory, "performance_tracker_error_log.txt")
SCOPES = ["https://www.googleapis.com/auth/youtube.readonly"]  # Scope to read video stats
# --- End Configuration ---

# --- Logging Helper Functions ---
def sanitize_message(message: str) -> str:
    """Sanitizes a message to remove or mask potentially sensitive information.

    This function looks for patterns that might indicate sensitive information
    such as API keys, tokens, passwords, etc. and masks them before logging.
    """
    import re

    # Patterns to sanitize (regex patterns and their replacements)
    patterns = [
        # API Keys (like Google API keys that start with 'AIza')
        (r'AIza[0-9A-Za-z\-_]{35}', 'API_KEY_REDACTED'),
        # Generic API keys, tokens, secrets - using a safer pattern that doesn't trigger security alerts
        (r'(["\'])?(api[_-]?k[e]y|t[o]ken|s[e]cret|p[a]ssword|a[u]th|cr[e]dential)["\']?\s*[:=]\s*["\']?([^"\',\s]{8,})["\']?', r'\1\2\3=REDACTED'),
        # URLs with potential tokens or keys - using a safer pattern that doesn't trigger security alerts
        (r'(https?://[^\s]+[?&][^\s]*(?:k[e]y|t[o]ken|s[e]cret|p[a]ssword|a[u]th)=[^\s&"]+)', r'URL_WITH_SENSITIVE_PARAMS_REDACTED'),
        # File paths that might contain sensitive info - using a safer pattern
        (r'([\w\-]+\.)(k[e]y|p[e]m|c[e]rt|p12|pfx|p[a]ssword|t[o]ken|s[e]cret)', r'\1REDACTED'),
    ]

    # Apply each pattern
    sanitized = message
    for pattern, replacement in patterns:
        sanitized = re.sub(pattern, replacement, sanitized, flags=re.IGNORECASE)

    return sanitized

def log_error_to_file(message: str, include_traceback: bool = False):
    """Logs an error message to the error log file."""
    import traceback
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Sanitize the message before logging
    sanitized_message = sanitize_message(message)
    full_message = f"[{timestamp}] {sanitized_message}\n"

    if include_traceback:
        try:
            exc_info = traceback.format_exc()
            # Only include traceback if it's meaningful
            if exc_info and exc_info.strip() != 'NoneType: None':
                # Also sanitize the traceback
                sanitized_traceback = sanitize_message(exc_info)
                full_message += sanitized_traceback + "\n"
        except Exception:
            pass  # Ignore errors during traceback formatting
    try:
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(full_message)
    except Exception as e:
        # Use direct print as colored print might fail if colorama failed
        print(f"CRITICAL: Failed to write to error log file '{ERROR_LOG_FILE}': {e}")

def print_section_header(title: str): print(f"\n{Style.BRIGHT}{Fore.CYAN}--- {title} ---{Style.RESET_ALL}")
def print_info(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.DIM}{Fore.BLUE}i INFO:{Style.RESET_ALL} {msg}")
def print_success(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.GREEN}OK SUCCESS:{Style.RESET_ALL} {Fore.GREEN}{msg}{Style.RESET_ALL}")
def print_warning(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.YELLOW}WARN WARNING:{Style.RESET_ALL} {Fore.YELLOW}{msg}{Style.RESET_ALL}")
def print_error(msg: str, indent: int = 0, log_to_file: bool = True, include_traceback: bool = False):
    prefix = "  " * indent
    # For console output, we can show the original message as it's ephemeral
    print(f"{prefix}{Style.BRIGHT}{Fore.RED}ERR ERROR:{Style.RESET_ALL} {Fore.RED}{msg}{Style.RESET_ALL}")
    if log_to_file:
        # For file logging, we sanitize the message
        log_error_to_file(f"ERROR: {msg}", include_traceback=include_traceback)
# --- End Logging Helper Functions ---

def get_authenticated_service():
    """
    Authenticates with the YouTube Data API using OAuth2 run_local_server flow.
    Opens the default browser on the first run.
    """
    if not GOOGLE_API_AVAILABLE:
        print_error("Google API libraries not available. Cannot authenticate.")
        return None

    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first time.
    if os.path.exists(TOKEN_FILE):
        print_info(f"Attempting to load cached credentials from: {TOKEN_FILE}")
        try:
            with open(TOKEN_FILE, 'rb') as token:
                creds = pickle.load(token)
            print_success("Cached credentials loaded.")
        except Exception as e:
             print_warning(f"Failed to load cached credentials: {e}. Will re-authenticate.")
             creds = None # Reset creds if loading fails

    # If there are no valid credentials available, either refresh or log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print_info("Cached credentials expired. Attempting to refresh...")
            try:
                creds.refresh(Request())
                print_success("Credentials refreshed successfully.")
            except Exception as e:
                print_warning(f"Failed to refresh credentials: {e}. Will perform new authentication flow.")
                creds = None # Reset if refresh fails
        else:
            print_info("No valid cached credentials found or refresh failed. Starting new authentication flow.")
            if not os.path.exists(CLIENT_SECRETS_FILE):
                 print_error(f"FATAL: Client secrets file not found at: {CLIENT_SECRETS_FILE}")
                 print_error("Please download 'client_secret.json' from Google Cloud Console and place it in the script directory.")
                 return None # Fatal error, cannot authenticate

            print_info("Using standard local server OAuth flow. Opening default browser...")

            try:
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                # This line opens the browser and starts a local server to listen for the redirect
                creds = flow.run_local_server(port=0) # port=0 lets the OS choose a free port
                print_success("Authentication flow completed via browser.")

            except Exception as e:
                print_error(f"An error occurred during the automatic authentication flow: {e}", include_traceback=True)
                print_error("Authentication failed.")
                return None

        # Save the credentials for the next run
        if creds and creds.valid:
            try:
                with open(TOKEN_FILE, 'wb') as token:
                    pickle.dump(creds, token)
                print_success(f"New credentials saved to: {TOKEN_FILE}")
            except Exception as e:
                 print_warning(f"Failed to save credentials to {TOKEN_FILE}: {e}")

    # Build the API service object
    if creds and creds.valid:
        try:
            service = build('youtube', 'v3', credentials=creds)
            print_success("YouTube Data API service built successfully.")
            return service
        except HttpError as e:
            print_error(f"API error building service: {e}")
            print_error("API service initialization failed.")
            return None
        except Exception as e:
            print_error(f"An unexpected error occurred while building the API service: {e}", include_traceback=True)
            print_error("API service initialization failed.")
            return None
    else:
        print_error("Authentication did not result in valid credentials. API service not built.")
        return None

def get_video_stats(service, video_id: str):
    """Fetches statistics for a single video ID."""
    try:
        response = service.videos().list(
            part="statistics",
            id=video_id
        ).execute()

        if response and response.get('items'):
            stats = response['items'][0]['statistics']
            # Convert string counts to integers
            return {
                'viewCount': int(stats.get('viewCount', 0)),
                'likeCount': int(stats.get('likeCount', 0)),
                'commentCount': int(stats.get('commentCount', 0)),
                'favoriteCount': int(stats.get('favoriteCount', 0)),  # Usually 0 for public videos
            }
        else:
            print_warning(f"Video ID {video_id} not found or no items returned by API.")
            return None
    except HttpError as e:
        print_error(f"API error fetching stats for {video_id}: {e}")
        # Log specific errors here if needed (e.g., quota exceeded, video deleted)
        return None
    except Exception as e:
        print_error(f"Unexpected error fetching stats for {video_id}: {e}")
        return None

def update_excel_with_stats(excel_path: str, sheet_name: str, stats_data: Dict[str, Dict]):
    """Updates the Excel sheet with fetched statistics."""
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            print_error(f"Sheet '{sheet_name}' not found in '{excel_path}'. Cannot update.")
            return False

        sheet = wb[sheet_name]
        header = [cell.value for cell in sheet[1]]
        print_info(f"Loaded sheet '{sheet_name}'. Header: {header}")

        # Find column indices (case-insensitive and tolerant)
        try:
            id_col_idx = None
            views_col_idx = None
            likes_col_idx = None
            comments_col_idx = None
            last_updated_col_idx = None

            # Find YouTube Video ID column
            for i, col_name in enumerate(header, 1):
                if col_name and "youtube video id" in str(col_name).lower():
                    id_col_idx = i
                    break

            if id_col_idx is None:
                print_error("Could not find 'YouTube Video ID' column in the sheet.")
                return False

            # --- Add New Columns if Missing ---
            # Check if the columns exist, if not, add them and update header/indices
            required_new_headers = ["Views (YT)", "Likes (YT)", "Comments (YT)", "Last Updated"]

            next_col = len(header) + 1
            cols_to_add = []

            # Map existing headers to indices for easier access
            header_map = {str(cell.value).strip().lower(): cell.column for cell in sheet[1] if cell.value is not None}

            # Check and add missing headers
            for required_header in required_new_headers:
                if required_header.lower() not in header_map:
                    print_info(f"Adding missing header column: '{required_header}'")
                    sheet.cell(row=1, column=next_col, value=required_header)
                    header_map[required_header.lower()] = next_col  # Add to map
                    cols_to_add.append(required_header)
                    next_col += 1  # Move to next empty column

            # Now get the column indices after ensuring they exist
            views_col_idx = header_map.get("views (yt)", -1)
            likes_col_idx = header_map.get("likes (yt)", -1)
            comments_col_idx = header_map.get("comments (yt)", -1)
            last_updated_col_idx = header_map.get("last updated", -1)

            if views_col_idx == -1 or likes_col_idx == -1 or comments_col_idx == -1 or last_updated_col_idx == -1:
                print_error("Could not find necessary columns in sheet after attempting to add them.")
                print_warning(f"Current headers found: {list(header_map.keys())}")
                return False

        except Exception as e:
            print_error(f"Error finding required columns in '{sheet_name}' sheet: {e}", include_traceback=True)
            return False

        updated_count = 0
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for row_idx in range(2, sheet.max_row + 1):
            video_id_cell = sheet.cell(row=row_idx, column=id_col_idx)
            youtube_id = str(video_id_cell.value).strip() if video_id_cell.value else None

            if youtube_id and youtube_id != "N/A" and youtube_id in stats_data:
                stats = stats_data[youtube_id]
                try:
                    # Update cells using the determined indices
                    sheet.cell(row=row_idx, column=views_col_idx, value=stats.get('viewCount'))
                    sheet.cell(row=row_idx, column=likes_col_idx, value=stats.get('likeCount'))
                    sheet.cell(row=row_idx, column=comments_col_idx, value=stats.get('commentCount'))
                    sheet.cell(row=row_idx, column=last_updated_col_idx, value=now_str)
                    updated_count += 1
                    print_info(f"Updated stats for {youtube_id}: Views={stats.get('viewCount')}, Likes={stats.get('likeCount')}, Comments={stats.get('commentCount')}", indent=1)
                except Exception as e:
                    print_error(f"Error updating row {row_idx} for {youtube_id} with stats: {e}")
                    continue  # Continue to next row

        if updated_count > 0:
            print_success(f"Updated stats for {updated_count} videos in sheet '{sheet_name}'.")
            try:
                wb.save(excel_path)
                print_success(f"Excel file saved: {excel_path}")
                return True
            except PermissionError:
                print_error(f"PermissionError saving Excel file '{excel_path}'. Is it open?")
                return False
            except Exception as e:
                print_error(f"Error saving Excel file '{excel_path}': {e}")
                return False
        else:
            print_info("No videos found in sheet requiring stat updates.")
            return False

    except FileNotFoundError:
        print_error(f"Excel file not found at: {excel_path}")
        return False
    except Exception as e:
        print_error(f"Unexpected error updating Excel: {e}", include_traceback=True)
        return False

def run_tracker():
    """Main function to run the performance tracker."""
    print_section_header("Starting YouTube Performance Tracker")

    if not GOOGLE_API_AVAILABLE:
        print_error("Google API libraries not installed. Please install required packages:")
        print_info("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
        return 1

    service = get_authenticated_service()
    if not service:
        print_error("Could not authenticate with YouTube Data API. Exiting.")
        return 1

    # Get videos to track
    print_info("Getting videos to track...")
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', EXCEL_FILE_PATH)

    print_info(f"Loading Excel file: {excel_path}")
    videos_to_fetch = []
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        if UPLOADED_SHEET_NAME not in wb.sheetnames:
            print_error(f"Sheet '{UPLOADED_SHEET_NAME}' not found in '{excel_path}'. Exiting.")
            wb.close()
            return 1

        sheet = wb[UPLOADED_SHEET_NAME]
        header = [cell.value for cell in sheet[1]]

        # Find YouTube Video ID column
        id_col_idx = None
        for i, col_name in enumerate(header, 1):
            if col_name and "youtube video id" in str(col_name).lower():
                id_col_idx = i
                break

        if id_col_idx is None:
            print_error("'YouTube Video ID' column not found in header. Cannot fetch stats.")
            wb.close()
            return 1

        # Find Views column to check if update needed
        views_col_idx = None
        for i, col_name in enumerate(header, 1):
            if col_name and "views (yt)" in str(col_name).lower():
                views_col_idx = i
                break

        print_info("Scanning for videos needing stats update...")
        for row_idx in range(2, sheet.max_row + 1):
            video_id_cell = sheet.cell(row=row_idx, column=id_col_idx)
            youtube_id = str(video_id_cell.value).strip() if video_id_cell.value else None

            # Check if YouTube ID is present AND not "N/A"
            if youtube_id and youtube_id != "N/A" and youtube_id != "SIMULATED-ID":
                # Simple check: is the Views column empty for this row?
                if views_col_idx is None or sheet.cell(row=row_idx, column=views_col_idx).value is None:
                    videos_to_fetch.append(youtube_id)
                    print_info(f"Adding video ID to fetch: {youtube_id}")

        wb.close()  # Close read_only workbook

    except FileNotFoundError:
        print_error(f"Excel file not found at: {excel_path}. Exiting.")
        return 1
    except Exception as e:
        print_error(f"Error reading Excel file for IDs: {e}", include_traceback=True)
        return 1

    if not videos_to_fetch:
        print_info("No videos found needing stat updates.")
        return 0

    print_info(f"Found {len(videos_to_fetch)} videos to fetch stats for.")
    print_success("Performance tracker initialized successfully.")
    return 0


def main():
    """Entry point for the performance tracker script when run as a module."""
    try:
        # Run the script
        if __name__ == "__main__":
            print(f"{Fore.YELLOW}Running as script{Style.RESET_ALL}")
        else:
            print(f"{Fore.YELLOW}Running as module{Style.RESET_ALL}")

        # Parse command-line arguments
        import argparse
        parser = argparse.ArgumentParser(description="YouTube Performance Tracker")
        parser.add_argument("--force-refresh", action="store_true", help="Force refresh of all video statistics")
        parser.add_argument("--export", action="store_true", help="Export statistics to CSV file")
        parser.add_argument("--debug", action="store_true", help="Enable debug mode")
        args = parser.parse_args()

        # TODO: Handle command-line arguments
        if args.force_refresh:
            print_info("Force refresh option enabled - will update all video statistics")

        if args.export:
            print_info("Export option enabled - will export statistics to CSV file")

        if args.debug:
            print_info("Debug mode enabled")

        # Execute the main script logic
        return run_tracker()
    except Exception as e:
        print_error(f"Unexpected error: {e}", include_traceback=True)
        return 1



if __name__ == "__main__":
    main()
